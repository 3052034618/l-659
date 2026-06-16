import os
from datetime import datetime, date, timedelta
from typing import Optional, Dict, List, Tuple
from sqlalchemy.orm import Session
from sqlalchemy import and_

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak,
)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from app.models import ComplianceReport, PermissionDeviation, AuditTicket
from app.core.config import settings
from app.utils import (
    logger,
    log_audit,
    get_deviation_type_text,
    get_risk_level_text,
    get_status_text,
    get_system_name,
)
from app.services.deviation_service import DeviationDetectionService


class ReportService:
    @classmethod
    def _register_fonts(cls):
        font_paths = [
            "C:/Windows/Fonts/msyh.ttc",
            "C:/Windows/Fonts/simhei.ttf",
            "C:/Windows/Fonts/simsun.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
            "/System/Library/Fonts/PingFang.ttc",
        ]
        registered = False
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont("ChineseFont", font_path))
                    registered = True
                    break
                except Exception:
                    continue
        if not registered:
            logger.warning("未找到中文字体，PDF中文可能显示异常")

    @classmethod
    def generate_daily_report(
        cls,
        db: Session,
        report_date: Optional[date] = None,
        operator_name: str = "system",
    ) -> ComplianceReport:
        report_date = report_date or date.today()

        existing = db.query(ComplianceReport).filter(
            ComplianceReport.report_date == report_date
        ).first()
        if existing:
            logger.info(f"{report_date} 合规报告已存在，跳过生成")
            return existing

        start_dt = datetime.combine(report_date - timedelta(days=30), datetime.min.time())
        end_dt = datetime.combine(report_date, datetime.max.time())

        stats = DeviationDetectionService.get_deviation_statistics(db, start_dt, end_dt)

        deviations_query = db.query(PermissionDeviation).filter(
            and_(
                PermissionDeviation.created_at >= start_dt,
                PermissionDeviation.created_at <= end_dt,
            )
        )
        deviations = deviations_query.all()

        system_stats = cls._calc_system_stats(db, deviations)

        total = stats["total"]
        by_risk = stats["by_risk"]
        by_status = stats["by_status"]

        report = ComplianceReport(
            report_date=report_date,
            report_type="daily",
            total_deviations=total,
            high_risk_count=by_risk.get("high", 0),
            medium_risk_count=by_risk.get("medium", 0),
            low_risk_count=by_risk.get("low", 0),
            resolved_count=by_status.get("resolved", 0),
            pending_count=by_status.get("pending", 0) + by_status.get("processing", 0),
            avg_fix_hours=stats["avg_fix_hours"],
            audit_completion_rate=stats["audit_completion_rate"],
            system_stats=system_stats,
            created_at=datetime.now(),
        )
        db.add(report)
        db.commit()
        db.refresh(report)

        try:
            pdf_path = cls.export_pdf(db, report)
            report.pdf_path = pdf_path
        except Exception as e:
            logger.error(f"生成PDF报告失败: {str(e)}")

        try:
            excel_path = cls.export_excel(db, report, deviations)
            report.excel_path = excel_path
        except Exception as e:
            logger.error(f"生成Excel报告失败: {str(e)}")

        db.commit()
        db.refresh(report)

        log_audit(
            db=db,
            action="generate_compliance_report",
            action_type="report",
            target_type="report",
            target_id=report.id,
            details=(
                f"生成{report_date}权限合规报告，"
                f"共{total}项偏离，"
                f"高危{report.high_risk_count}项，"
                f"完成率{report.audit_completion_rate}%"
            ),
            username=operator_name,
            status="success",
        )

        logger.info(
            f"{report_date}合规报告生成完成: 偏离{total}项，"
            f"完成率{report.audit_completion_rate}%"
        )
        return report

    @classmethod
    def _calc_system_stats(
        cls,
        db: Session,
        deviations: List[PermissionDeviation],
    ) -> Dict:
        stats: Dict[str, Dict] = {}
        for dev in deviations:
            sys_code = dev.system_code
            if sys_code not in stats:
                stats[sys_code] = {
                    "system_name": get_system_name(sys_code),
                    "total": 0,
                    "high": 0,
                    "medium": 0,
                    "low": 0,
                    "resolved": 0,
                    "excessive": 0,
                    "deficient": 0,
                }
            stats[sys_code]["total"] += 1
            stats[sys_code][dev.risk_level] = stats[sys_code].get(dev.risk_level, 0) + 1
            if dev.status == "resolved":
                stats[sys_code]["resolved"] += 1
            stats[sys_code][dev.deviation_type] = stats[sys_code].get(dev.deviation_type, 0) + 1

        return stats

    @classmethod
    def export_excel(
        cls,
        db: Session,
        report: ComplianceReport,
        deviations: List[PermissionDeviation],
    ) -> str:
        filename = f"权限合规报告_{report.report_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(settings.EXPORT_DIR, filename)

        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "报告概览"
        cls._write_excel_summary(ws_summary, report)

        ws_deviations = wb.create_sheet("权限偏离明细")
        cls._write_excel_deviations(ws_deviations, deviations)

        ws_systems = wb.create_sheet("各系统统计")
        cls._write_excel_system_stats(ws_systems, report.system_stats or {})

        wb.save(filepath)
        logger.info(f"Excel报告已保存: {filepath}")
        return filepath

    @classmethod
    def _write_excel_summary(cls, ws, report: ComplianceReport):
        title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="微软雅黑", size=11)
        title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = f"权限合规审计日报 - {report.report_date.strftime('%Y年%m月%d日')}"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center_align
        ws.row_dimensions[1].height = 40

        summary_data = [
            ["指标", "数值", "指标", "数值", "指标", "数值"],
            ["权限偏离总数", report.total_deviations, "高危偏离数", report.high_risk_count,
             "中危偏离数", report.medium_risk_count],
            ["低危偏离数", report.low_risk_count, "待处理数", report.pending_count,
             "已解决数", report.resolved_count],
            ["平均修复时长(小时)", report.avg_fix_hours,
             "审计完成率(%)", report.audit_completion_rate, "", ""],
        ]

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20

    @classmethod
    def _write_excel_deviations(cls, ws, deviations: List[PermissionDeviation]):
        headers = [
            "ID", "系统", "权限名称", "偏离类型", "风险等级",
            "风险分值", "状态", "描述", "创建时间", "解决时间",
        ]
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        data_font = Font(name="微软雅黑", size=10)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        risk_fills = {
            "high": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
            "low": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        }

        for row_idx, dev in enumerate(deviations, start=2):
            row_data = [
                dev.id,
                get_system_name(dev.system_code),
                dev.permission_name,
                get_deviation_type_text(dev.deviation_type),
                get_risk_level_text(dev.risk_level),
                dev.risk_score,
                get_status_text(dev.status),
                dev.description or "",
                dev.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                dev.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if dev.resolved_at else "",
            ]
            fill = risk_fills.get(dev.risk_level)

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border
                if fill and col_idx == 5:
                    cell.fill = fill

        col_widths = [8, 15, 25, 25, 10, 10, 10, 50, 20, 20]
        for idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width

    @classmethod
    def _write_excel_system_stats(cls, ws, system_stats: Dict):
        headers = [
            "系统代码", "系统名称", "偏离总数", "高危", "中危",
            "低危", "已解决", "权限过高", "权限过低",
        ]
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        data_font = Font(name="微软雅黑", size=10)
        center_align = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, (sys_code, stats) in enumerate(system_stats.items(), start=2):
            row_data = [
                sys_code,
                stats.get("system_name", sys_code),
                stats.get("total", 0),
                stats.get("high", 0),
                stats.get("medium", 0),
                stats.get("low", 0),
                stats.get("resolved", 0),
                stats.get("excessive", 0),
                stats.get("deficient", 0),
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = center_align
                cell.border = thin_border

        for idx in range(1, 10):
            ws.column_dimensions[get_column_letter(idx)].width = 15

    @classmethod
    def export_pdf(
        cls,
        db: Session,
        report: ComplianceReport,
    ) -> str:
        cls._register_fonts()

        filename = f"权限合规报告_{report.report_date.strftime('%Y%m%d')}.pdf"
        filepath = os.path.join(settings.EXPORT_DIR, filename)

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            rightMargin=2 * cm, leftMargin=2 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        font_name = "ChineseFont" if "ChineseFont" in pdfmetrics.getRegisteredFontNames() else "Helvetica"

        title_style = ParagraphStyle(
            "TitleChinese", parent=styles["Title"],
            fontName=font_name, fontSize=20, alignment=1, spaceAfter=20,
        )
        h2_style = ParagraphStyle(
            "H2Chinese", parent=styles["Heading2"],
            fontName=font_name, fontSize=14, spaceBefore=15, spaceAfter=10,
        )
        normal_style = ParagraphStyle(
            "NormalChinese", parent=styles["Normal"],
            fontName=font_name, fontSize=10, leading=14,
        )

        elements = []
        elements.append(Paragraph(
            f"权限合规审计日报<br/>{report.report_date.strftime('%Y年%m月%d日')}",
            title_style,
        ))

        elements.append(Paragraph("一、总体概览", h2_style))
        summary_data = [
            ["指标", "数值", "指标", "数值"],
            ["权限偏离总数", str(report.total_deviations), "高危偏离数", str(report.high_risk_count)],
            ["中危偏离数", str(report.medium_risk_count), "低危偏离数", str(report.low_risk_count)],
            ["待处理数", str(report.pending_count), "已解决数", str(report.resolved_count)],
            ["平均修复时长(小时)", str(report.avg_fix_hours), "审计完成率(%)", f"{report.audit_completion_rate}%"],
        ]
        summary_table = Table(summary_data, colWidths=[4 * cm, 3 * cm, 4 * cm, 3 * cm])
        summary_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]))
        elements.append(summary_table)

        elements.append(Spacer(1, 0.5 * cm))
        elements.append(Paragraph("二、各系统偏离统计", h2_style))

        sys_headers = ["系统", "总数", "高危", "中危", "低危", "已解决", "完成率"]
        sys_table_data = [sys_headers]
        system_stats = report.system_stats or {}
        for sys_code, stats in system_stats.items():
            total = stats.get("total", 0)
            resolved = stats.get("resolved", 0)
            rate = f"{round(resolved / total * 100, 1)}%" if total > 0 else "0%"
            sys_table_data.append([
                stats.get("system_name", sys_code),
                str(total),
                str(stats.get("high", 0)),
                str(stats.get("medium", 0)),
                str(stats.get("low", 0)),
                str(resolved),
                rate,
            ])

        if len(sys_table_data) == 1:
            sys_table_data.append(["暂无数据", "-", "-", "-", "-", "-", "-"])

        sys_table = Table(sys_table_data, colWidths=[3.5 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 1.5 * cm, 2 * cm])
        sys_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#5B9BD5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]))
        elements.append(sys_table)

        elements.append(Spacer(1, 0.5 * cm))
        elements.append(Paragraph("三、高危偏离清单", h2_style))

        high_risk = db.query(PermissionDeviation).filter(
            and_(
                PermissionDeviation.risk_level == "high",
                PermissionDeviation.status.in_(["pending", "processing"]),
            )
        ).limit(20).all()

        risk_headers = ["系统", "权限", "偏离类型", "状态", "创建时间"]
        risk_table_data = [risk_headers]
        for dev in high_risk:
            risk_table_data.append([
                get_system_name(dev.system_code),
                dev.permission_name[:20],
                get_deviation_type_text(dev.deviation_type),
                get_status_text(dev.status),
                dev.created_at.strftime("%m-%d %H:%M"),
            ])

        if len(risk_table_data) == 1:
            risk_table_data.append(["暂无", "未处理的", "高危", "偏离项", ""])

        risk_table = Table(risk_table_data, colWidths=[2.5 * cm, 4.5 * cm, 3 * cm, 2 * cm, 2.5 * cm])
        risk_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FF6B6B")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ]))
        elements.append(risk_table)

        elements.append(Spacer(1, 1 * cm))
        elements.append(Paragraph(
            f"<br/>报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            normal_style,
        ))
        elements.append(Paragraph(
            f"系统：{settings.APP_NAME} v{settings.APP_VERSION}",
            normal_style,
        ))

        doc.build(elements)
        logger.info(f"PDF报告已保存: {filepath}")
        return filepath

    @classmethod
    def get_reports(
        cls,
        db: Session,
        report_type: Optional[str] = None,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        page: int = 1,
        page_size: int = 20,
    ) -> Tuple[List[ComplianceReport], int]:
        query = db.query(ComplianceReport)
        if report_type:
            query = query.filter(ComplianceReport.report_type == report_type)
        if start_date:
            query = query.filter(ComplianceReport.report_date >= start_date)
        if end_date:
            query = query.filter(ComplianceReport.report_date <= end_date)

        total = query.count()
        reports = (
            query.order_by(ComplianceReport.report_date.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return reports, total
